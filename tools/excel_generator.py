import os
import uuid
import json
import re
import sqlite3
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from open_webui.models.files import Files, FileForm
except ImportError:
    from open_webui.apps.webui.models.files import Files, FileForm


class Tools:
    def __init__(self):
        pass

    # ============================================================
    # INSTANT KB LINKER
    # ============================================================

    def _link_to_output_kb(self, file_id: str, user_id: str) -> None:
        """
        Instantly links the generated file to the 'output' Knowledge Base in SQLite.
        """
        db_path = "./data/webui.db"
        try:
            con = sqlite3.connect(db_path)
            cur = con.cursor()

            # 1. Fetch or create 'output' KB ID
            cur.execute("SELECT id FROM knowledge WHERE name='output'")
            row = cur.fetchone()

            if row:
                kb_id = row[0]
            else:
                kb_id = str(uuid.uuid4())
                now = int(time.time())
                cur.execute(
                    "INSERT INTO knowledge (id, user_id, name, description, created_at, updated_at) "
                    "VALUES (?, ?, 'output', 'Auto-collected OpenWebUI-generated files', ?, ?)",
                    (kb_id, user_id, now, now),
                )

            # 2. Insert link row into knowledge_file table instantly
            now = int(time.time())
            cur.execute(
                "INSERT OR IGNORE INTO knowledge_file "
                "(id, user_id, knowledge_id, file_id, created_at, updated_at, directory_id) "
                "VALUES (?, ?, ?, ?, ?, ?, NULL)",
                (str(uuid.uuid4()), user_id, kb_id, file_id, now, now),
            )
            con.commit()
            con.close()
        except Exception as e:
            print(f"Failed to instantly link file to output KB: {e}")

    # ============================================================
    # Utility Functions
    # ============================================================

    def _clean_filename(self, filename: str) -> str:

        if not filename:
            filename = "generated_excel"

        filename = filename.strip()

        filename = re.sub(r'[<>:"/\\|?*]', "_", filename)

        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        return filename

    def _clean_sheet_name(self, name: str, existing_names=None):

        if existing_names is None:
            existing_names = []

        if not name:
            name = "Sheet1"

        name = str(name).strip()

        # Excel forbidden characters
        name = re.sub(r"[\[\]\*:/\\?]", "_", name)

        name = name[:31]

        original = name
        counter = 1

        while name in existing_names:

            suffix = f"_{counter}"

            name = original[: 31 - len(suffix)] + suffix

            counter += 1

        return name

    # ============================================================
    # Professional Excel Formatting
    # ============================================================

    def _format_sheet(self, ws, freeze_header=True, add_filter=True):

        if ws.max_row < 1:
            return

        # --------------------------------------------------------
        # Header
        # --------------------------------------------------------

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        header_font = Font(bold=True, color="FFFFFF")

        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        border = Border(bottom=Side(style="thin", color="D9E1F2"))

        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        ws.row_dimensions[1].height = 30

        # --------------------------------------------------------
        # Freeze header
        # --------------------------------------------------------

        if freeze_header:
            ws.freeze_panes = "A2"

        # --------------------------------------------------------
        # Auto filter
        # --------------------------------------------------------

        if add_filter and ws.max_row >= 2:

            ws.auto_filter.ref = ws.dimensions

        # --------------------------------------------------------
        # Cell formatting
        # --------------------------------------------------------

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # --------------------------------------------------------
        # Column width
        # --------------------------------------------------------

        for column_cells in ws.columns:

            if not column_cells:
                continue

            column_number = column_cells[0].column

            max_length = 0

            for cell in column_cells:

                try:

                    if cell.value is None:
                        value = ""

                    else:
                        value = str(cell.value)

                    # Limit very long text
                    max_length = max(max_length, min(len(value), 50))

                except Exception:
                    pass

            width = max(12, min(max_length + 3, 40))

            ws.column_dimensions[get_column_letter(column_number)].width = width

    # ============================================================
    # Create Excel Table
    # ============================================================

    def _add_excel_table(self, ws):

        if ws.max_row < 2:
            return

        if ws.max_column < 1:
            return

        reference = f"A1:" f"{get_column_letter(ws.max_column)}" f"{ws.max_row}"

        # Excel table names must be unique
        table_name = "Table_" + re.sub(r"[^A-Za-z0-9_]", "", ws.title)

        if not table_name:
            table_name = "Table_Data"

        table_name = table_name[:200]

        try:

            table = Table(displayName=table_name, ref=reference)

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style

            ws.add_table(table)

        except Exception:
            # Formatting failure should not
            # prevent Excel generation.
            pass

    # ============================================================
    # Write Structured Data
    # ============================================================

    def _write_structured_sheet(self, ws, columns, rows):

        # --------------------------------------------------------
        # Validate columns
        # --------------------------------------------------------

        if not columns:

            columns = ["Data"]

        columns = [str(column).strip() for column in columns]

        # Remove empty column names
        columns = [column for column in columns if column]

        if not columns:

            columns = ["Data"]

        # --------------------------------------------------------
        # Create header row
        # --------------------------------------------------------

        for column_index, column_name in enumerate(columns, start=1):

            ws.cell(row=1, column=column_index, value=column_name)

        # --------------------------------------------------------
        # Create rows
        # --------------------------------------------------------

        if rows is None:
            rows = []

        for row_index, row in enumerate(rows, start=2):

            # ----------------------------------------------------
            # Dictionary row
            # ----------------------------------------------------

            if isinstance(row, dict):

                for column_index, column_name in enumerate(columns, start=1):

                    value = row.get(column_name, "")

                    ws.cell(row=row_index, column=column_index, value=value)

            # ----------------------------------------------------
            # List row
            # ----------------------------------------------------

            elif isinstance(row, list):

                for column_index in range(1, len(columns) + 1):

                    if column_index <= len(row):

                        value = row[column_index - 1]

                    else:

                        value = ""

                    ws.cell(row=row_index, column=column_index, value=value)

            # ----------------------------------------------------
            # Single value
            # ----------------------------------------------------

            else:

                ws.cell(row=row_index, column=1, value=row)

        # --------------------------------------------------------
        # Formatting
        # --------------------------------------------------------

        self._format_sheet(ws)

        self._add_excel_table(ws)

    # ============================================================
    # Create Excel File
    # ============================================================

    async def create_excel_file(
        self,
        workbook_data: str,
        filename: str = "generated_excel.xlsx",
        __user__: dict = None,
        __messages__: list = None,
    ) -> dict:
        """
        Create a professional Excel workbook.

        IMPORTANT:
        Before calling this tool, determine the correct
        columns and rows from the user's request.

        The workbook_data parameter must contain JSON.

        Example:

        {
            "sheets": [
                {
                    "name": "Project Inputs",
                    "columns": [
                        "Section",
                        "Parameter",
                        "Value",
                        "Unit"
                    ],
                    "rows": [
                        {
                            "Section": "Project Inputs",
                            "Parameter": "Initial Investment",
                            "Value": 5000000,
                            "Unit": "INR"
                        }
                    ]
                }
            ]
        }

        The model should decide the appropriate columns
        and rows based on the user's request.

        For "convert this chat into Excel", use the
        conversation messages to determine a meaningful
        tabular structure rather than simply putting
        every message into one column.
        """

        # ========================================================
        # Parse workbook JSON
        # ========================================================

        if not workbook_data:
            return {"error": "No workbook data was provided."}

        try:
            workbook_definition = json.loads(workbook_data)
        except Exception as e:
            return {
                "error": f"workbook_data must be valid JSON. Details: {str(e)}"
            }

        # ========================================================
        # Validate workbook structure
        # ========================================================

        if not isinstance(workbook_definition, dict):
            return {"error": "Workbook definition must be a JSON object."}

        sheets = workbook_definition.get("sheets", [])

        if not sheets:
            return {"error": "No sheets were defined."}

        # ========================================================
        # Create Workbook
        # ========================================================

        workbook = Workbook()

        # Remove default sheet
        default_sheet = workbook.active

        workbook.remove(default_sheet)

        created_sheets = []

        # ========================================================
        # Create Each Sheet
        # ========================================================

        for sheet_definition in sheets:

            if not isinstance(sheet_definition, dict):
                continue

            sheet_name = sheet_definition.get("name", "Sheet")

            sheet_name = self._clean_sheet_name(sheet_name, workbook.sheetnames)

            columns = sheet_definition.get("columns", [])

            rows = sheet_definition.get("rows", [])

            ws = workbook.create_sheet(title=sheet_name)

            self._write_structured_sheet(ws, columns, rows)

            created_sheets.append(sheet_name)

        # ========================================================
        # Ensure Sheet Exists
        # ========================================================

        if not created_sheets:

            ws = workbook.create_sheet(title="Data")

            ws["A1"] = "No data"

            created_sheets.append("Data")

        # ========================================================
        # Create Information Sheet
        # ========================================================

        info_ws = workbook.create_sheet(
            title=self._clean_sheet_name("Information", workbook.sheetnames)
        )

        information = [
            ["Property", "Value"],
            ["Generated By", "Open WebUI Excel Generator"],
            ["Brand", "Spryntworks"],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Number of Sheets", len(created_sheets)],
            ["Sheets", ", ".join(created_sheets)],
        ]

        for row_index, row in enumerate(information, start=1):

            for column_index, value in enumerate(row, start=1):

                info_ws.cell(row=row_index, column=column_index, value=value)

        self._format_sheet(info_ws, freeze_header=False, add_filter=False)

        # ========================================================
        # Save File
        # ========================================================

        filename = self._clean_filename(filename)

        upload_dir = os.path.abspath("./data/uploads")

        os.makedirs(upload_dir, exist_ok=True)

        file_id = str(uuid.uuid4())

        unique_filename = f"{file_id}_{filename}"

        file_path = os.path.join(upload_dir, unique_filename)

        workbook.save(file_path)

        # ========================================================
        # Register with Open WebUI
        # ========================================================

        user_id = __user__.get("id") if __user__ else "admin"

        file_size = os.path.getsize(file_path)

        file_form = FileForm(
            id=file_id,
            filename=filename,
            path=file_path,
            data={"workbook_definition": workbook_data},
            meta={
                "name": filename,
                "content_type": "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet",
                "size": file_size,
                "description": "Excel workbook generated by " "Open WebUI",
                "tags": ["excel", "xlsx", "spreadsheet", "generated"],
            },
        )

        result = await Files.insert_new_file(user_id, file_form)

        if result is None:
            return {
                "error": "Excel file was created but could not be registered inside Open WebUI."
            }

        # INSTANTLY LINK TO 'output' KNOWLEDGE BASE
        self._link_to_output_kb(file_id, user_id)

        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Return structured output format for chat auto-attachment
        return {
            "output": (
                f"Excel file created successfully. "
                f"Filename: {filename}. Sheets: {', '.join(created_sheets)}"
            ),
            "files": [
                {
                    "id": file_id,
                    "filename": filename,
                    "name": filename,
                    "url": file_id,
                    "content_type": content_type,
                    "size": file_size,
                    "type": "file",
                    "status": "processed",
                }
            ],
        }
